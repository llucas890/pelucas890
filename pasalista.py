from operator import length_hint
import string
from tracemalloc import stop
from openpyxl import load_workbook
from openpyxl.xml.constants import MAX_ROW

FILE_PATH = 'E:\Documentos\inva\INVA2.xlsx'

workbook = load_workbook(FILE_PATH)
wb = workbook.active
nombresnoencontrados = []
listabusqueda = []
busqueda = ''
stock = wb['B17']
while busqueda != '6':
    print('')
    print('1 Para sumar asistencias')
    print('2 Para cobros')
    print('3 Para agregar nuevos usuarios')
    print('4 Para agregar una box al stock')
    print('5 Para guardar y salir')
    print('6 Para salir sin guardar')
    print('7 Para eliminar un usuario de la lista')
    print('8 Para restar asist a todos')
    print('9 Eliminar asistencias')
    print('Programa creado por Lucas Lescano alias Desxz')
    print('https://www.linkedin.com/in/lucaslescano20/')
    busqueda = input('Ingresa la opcion: ')
    print('')
    if busqueda == '1':
        while busqueda != 'SALIR':
                busqueda = input('Introduce el nombre o 2 para lista de no encontrados: ')         
                busqueda = busqueda.upper()
                busqueda = busqueda.replace(" ", "")
                busqueda = busqueda.replace("_", "")
                busqueda = busqueda.replace("-", "")
                busqueda = busqueda.replace("0", "o")
                busqueda = busqueda.replace("SrMeLiОDaS", "MELIO")
                busqueda = busqueda.replace(".", "")
                busqueda = busqueda.replace("sm", "")
                busqueda = busqueda.replace("(box)", "")
                busqueda = busqueda.replace("box", "")
                busqueda = busqueda.replace("BK", "")
                listabusqueda.append(busqueda)
                if len(listabusqueda) != len(set(listabusqueda)):
                    print(listabusqueda)
                    print(f"Se encontro un duplicado {busqueda}")
                    listabusqueda.clear()
                    break
                for colA,colNB,colC, colND,colE, colNF,colG, colNH, colI, colNJ in wb:
                    if busqueda in colA.value:
                        wb[f'B{colNB.row}'] = colNB.value + 1
                        print(f"Nombre encontrado: {busqueda}. Asistencias: {wb[f'B{colNB.row}'].value}\n")
                        break
                    elif busqueda in colC.value:
                        wb[f'D{colND.row}'] = colND.value + 1
                        print(f"Nombre encontrado: {busqueda}. Asistencias: {wb[f'D{colND.row}'].value}\n")
                        break
                    elif busqueda in colE.value:
                        wb[f'F{colNF.row}'] = colNF.value + 1
                        print(f"Nombre encontrado: {busqueda}. Asistencias: {wb[f'F{colNF.row}'].value}\n")
                        break
                    elif busqueda in colG.value:
                        wb[f'H{colNH.row}'] = colNH.value + 1
                        print(f"Nombre encontrado: {busqueda}. Asistencias: {wb[f'H{colNH.row}'].value}\n")
                        break
                    elif busqueda in colI.value:
                        wb[f'J{colNJ.row}'] = colNJ.value + 1
                        print(f"Nombre encontrado: {busqueda}. Asistencias: {wb[f'J{colNJ.row}'].value}\n")
                        break
                else:
                    nombresnoencontrados.append(busqueda)
                if busqueda == '2':
                    print(f"Nombres no encontrados: {nombresnoencontrados}")
                
    if busqueda == '4':
        busqueda = 'STOCK'
        for colA,colNB,colC, colND,colE, colNF,colG, colNH, colI, colNJ in wb:
            if busqueda in colA.value:
                wb[f'B{colNB.row}'] = colNB.value + 1
                print(f"Se sumo una box al stock: {wb[f'B{colNB.row}'].value}\n")
    if busqueda == '5':       
        workbook.save('E:\Documentos\inva\Inva2.xlsx')
        print('Guardado Exitoso'.center(40, "="))
        break
    if busqueda == '8':
        numero = input("Cuantas asist deseas restar?: ")
        numero = int(numero)
        for row in range(2, wb.max_row+1):
            cell = wb.cell(row=row, column=2)
            if cell.value is not wb['B17']:
                cell.value = cell.value - numero
        for row in range(2, wb.max_row+1):
            cell = wb.cell(row=row, column=4)
            if cell.value is not wb['B17']:
                cell.value = cell.value - numero
        for row in range(2, wb.max_row+1):
            cell = wb.cell(row=row, column=6)
            if cell.value is not wb['B17']:
                cell.value = cell.value - numero
        for row in range(2, wb.max_row+1):
            cell = wb.cell(row=row, column=8)
            if cell.value is not wb['B17']:
                cell.value = cell.value - numero
        for row in range(2, wb.max_row+1):
            cell = wb.cell(row=row, column=10)
            if cell.value is not wb['B17']:
                cell.value = cell.value - numero
        stock.value = stock.value + numero
        print(f"Se restaron correctamente las {numero} asistencias a todos.")
        print(f"Stock: {stock.value}")
    if busqueda == '9':
        while busqueda != 'SALIR':
                busqueda = input('////// Introduce A RESTAR: /////')         
                busqueda = busqueda.upper()
                busqueda = busqueda.strip()
                busqueda = busqueda.replace(" ", "")
                busqueda = busqueda.replace("-", "")
                busqueda = busqueda.replace("0", "o")
                busqueda = busqueda.replace("SrMeLiОDaS", "MELIO")
                busqueda = busqueda.replace(".", "")
                for colA,colNB,colC, colND,colE, colNF,colG, colNH, colI, colNJ in wb:
                    if busqueda in colA.value:
                        wb[f'B{colNB.row}'] = colNB.value - 1
                        print(f"\nNombre encontrado: {busqueda}. Asistencias: {wb[f'B{colNB.row}'].value}\n")
                    elif busqueda in colC.value:
                        wb[f'D{colND.row}'] = colND.value - 1
                        print(f"\nNombre encontrado: {busqueda}. Asistencias: {wb[f'D{colND.row}'].value}\n")
                    elif busqueda in colE.value:
                        wb[f'F{colNF.row}'] = colNF.value - 1
                        print(f"\nNombre encontrado: {busqueda}. Asistencias: {wb[f'F{colNF.row}'].value}\n")
                    elif busqueda in colG.value:
                        wb[f'H{colNH.row}'] = colNH.value - 1
                        print(f"\nNombre encontrado: {busqueda}. Asistencias: {wb[f'H{colNH.row}'].value}\n")
                    elif busqueda in colI.value:
                        wb[f'J{colNJ.row}'] = colNJ.value - 1
                        print(f"\nNombre encontrado: {busqueda}. Asistencias: {wb[f'J{colNJ.row}'].value}\n")
    if busqueda == '2':
        busqueda = input('Ingrese el nombre de la persona que cobro: ')
        busqueda = busqueda.upper()
        box = input('Cuantas box cobro?: ')
        if box == "1":
            asistencias = int(20)
            for colA,colNB,colC, colND,colE, colNF,colG, colNH, colI, colNJ in wb:
                if busqueda in colA.value:
                    wb[f'B{colNB.row}'] = colNB.value - asistencias
                    print(f"COBRO: {busqueda} {box}+5. Asistencias: {wb[f'B{colNB.row}'].value}")
                    break
                elif busqueda in colC.value:
                    wb[f'D{colND.row}'] = colND.value - asistencias
                    print(f"COBRO: {busqueda} {box}+5. Asistencias: {wb[f'D{colND.row}'].value}")
                    break
                elif busqueda in colE.value:
                    wb[f'F{colNF.row}'] = colNF.value - asistencias
                    print(f"COBRO: {busqueda} {box}+5. Asistencias: {wb[f'F{colNF.row}'].value}")
                    break
                elif busqueda in colG.value:
                    wb[f'H{colNH.row}'] = colNH.value - asistencias
                    print(f"COBRO: {busqueda} {box}+5. Asistencias: {wb[f'H{colNH.row}'].value}")
                    break
                elif busqueda in colI.value:
                    wb[f'J{colNJ.row}'] = colNJ.value - asistencias
                    print(f"COBRO: {busqueda} {box}+5. Asistencias: {wb[f'J{colNJ.row}'].value}")
                    break     
            stock.value= stock.value - int(box)
            print(f"Se resto {box} box al stock: {stock.value}\n")
        elif box == "2":
            asistencias = int(34)
            for colA,colNB,colC, colND,colE, colNF,colG, colNH, colI, colNJ in wb:
                if busqueda in colA.value:
                    wb[f'B{colNB.row}'] = colNB.value - asistencias
                    print(f"COBRO: {busqueda} {box}+5. Asistencias: {wb[f'B{colNB.row}'].value}")
                    break
                elif busqueda in colC.value:
                    wb[f'D{colND.row}'] = colND.value - asistencias
                    print(f"COBRO: {busqueda} {box}+5. Asistencias: {wb[f'D{colND.row}'].value}")
                    break
                elif busqueda in colE.value:
                    wb[f'F{colNF.row}'] = colNF.value - asistencias
                    print(f"COBRO: {busqueda} {box}+5. Asistencias: {wb[f'F{colNF.row}'].value}")
                    break
                elif busqueda in colG.value:
                    wb[f'H{colNH.row}'] = colNH.value - asistencias
                    print(f"COBRO: {busqueda} {box}+5. Asistencias: {wb[f'H{colNH.row}'].value}")
                    break
                elif busqueda in colI.value:
                    wb[f'J{colNJ.row}'] = colNJ.value - asistencias
                    print(f"COBRO: {busqueda} {box}+5. Asistencias: {wb[f'J{colNJ.row}'].value}")
                    break     
            stock.value= stock.value - int(box)
            print(f"Se resto {box} box al stock: {stock.value}\n")
    if busqueda == "3":
        agregar = input("Ingrese el nuevo personaje: ")
        agregar = agregar.upper()
        busqueda = "LIBRE"
        for colA,colNB,colC, colND,colE, colNF,colG, colNH, colI, colNJ in wb:
            if busqueda == colA.value:
                wb[f'A{colA.row}'] = agregar
                wb[f'B{colNB.row}'] = 0
                print(f"Se agrego el usuario {agregar} en columna A")
                break
            elif busqueda == colC.value:
                wb[f'C{colC.row}'] = agregar
                wb[f'D{colND.row}'] = 0
                print(f"Se agrego el usuario {agregar} en columna C")
                break
            elif busqueda == colE.value:
                wb[f'E{colE.row}'] = agregar
                wb[f'F{colNF.row}'] = 0
                print(f"Se agrego el usuario {agregar} en columna E")
                break
            elif busqueda == colG.value:
                wb[f'G{colG.row}'] = agregar
                wb[f'H{colNH.row}'] = 0
                print(f"Se agrego el usuario {agregar} en columna G")
                break
            elif busqueda == colI.value:
                wb[f'I{colI.row}'] = agregar
                wb[f'J{colNJ.row}'] = 0
                print(f"Se agrego el usuario {agregar} en columna I")
                break
        else:
            print("No se encontro un slot libre")
    if busqueda == "7":
        agregar = input("Ingrese el usuario a eliminar: ")
        agregar = agregar.upper()
        busqueda = agregar
        for colA,colNB,colC, colND,colE, colNF,colG, colNH, colI, colNJ in wb:
            if busqueda == colA.value:
                wb[f'A{colA.row}'] = "LIBRE"
                wb[f'B{colNB.row}'] = 0
                print(f"Se elimino el usuario {agregar} en columna A")
                break
            elif busqueda == colC.value:
                wb[f'C{colC.row}'] = "LIBRE"
                wb[f'D{colND.row}'] = 0
                print(f"Se elimino el usuario {agregar} en columna C")
                break
            elif busqueda == colE.value:
                wb[f'E{colE.row}'] = "LIBRE"
                wb[f'F{colNF.row}'] = 0
                print(f"Se ELIMINO el usuario {agregar} en columna E")
                break
            elif busqueda == colG.value:
                wb[f'G{colG.row}'] = "LIBRE"
                wb[f'H{colNH.row}'] = 0
                print(f"Se ELIMINO el usuario {agregar} en columna G")
                break
            elif busqueda == colI.value:
                wb[f'I{colI.row}'] = "LIBRE"
                wb[f'J{colNJ.row}'] = 0
                print(f"Se ELIMINO el usuario {agregar} en columna I")
                break
        else:
            print("No se encontro el usuario a eliminar")



